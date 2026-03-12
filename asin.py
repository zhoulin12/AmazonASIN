import requests
from playwright.sync_api import sync_playwright
import json
import re
from datetime import datetime
import pandas as pd
from bs4 import BeautifulSoup
from io import BytesIO
try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage
except Exception:
    load_workbook = None
    XLImage = None
    get_column_letter = None
    PILImage = None
import html

def _split_image_urls(val):
    if not val:
        return []
    s = html.unescape(str(val)).strip()
    parts = re.split(r'[\r\n,;|\t ]+', s)
    urls = []
    for p in parts:
        u = p.strip()
        if not u:
            continue
        if u.startswith('http'):
            urls.append(u)
    return urls

def extract_details_from_html(html):
    """
    通用解析函数，从 HTML 中提取商品标题、图片、详情等。
    兼容 ProductData.py 的逻辑。
    """
    soup = BeautifulSoup(html, 'html.parser')
    details = {}
    
    # 商品标题
    title_tag = soup.find('span', id='productTitle')
    if not title_tag:
        title_tag = soup.find('span', class_='a-size-large product-title-word-break')
    details['商品标题'] = title_tag.get_text(strip=True) if title_tag else ""
    
    # 商品图片
    images = []
    alt_images_div = soup.find('div', id='altImages')
    if alt_images_div:
        img_tags = alt_images_div.find_all('img')
        images = [img.get('src') for img in img_tags if img.get('src')]
    details['商品图片'] = images
    
    # 商品详情
    detail = soup.find('div', id='poExpander')
    details['商品详情'] = detail.get_text(strip=True) if detail else ""
    
    # 商品详情2
    detail2 = soup.find('div', id='feature-bullets')
    details['商品详情2'] = detail2.get_text(strip=True) if detail2 else ""

    # 商品详情3
    detail3 = soup.find('div', id='prodDetails')
    details['商品详情3'] = detail3.get_text(strip=True) if detail3 else ""

    # 商品详情图片描述
    xq_texts = []
    aplus_div = soup.find('div', id='aplus')
    if aplus_div:
        p_tags = aplus_div.find_all('p')
        xq_texts = [p.get_text(strip=True) for p in p_tags]
    details['详情图片描述'] = "\n".join(xq_texts)

    # 商品详情图片
    xq_imgs = []
    if aplus_div:
        img_tags = aplus_div.find_all('img')
        xq_imgs = [img.get('src') for img in img_tags if img.get('src')]
    details['详情图片'] = xq_imgs
    
    return details

def capture_variants_only(product_url):
    """
    使用 Playwright 仅用于捕获 AJAX 响应（获取变体列表和价格）。
    """
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=300)
        context = browser.new_context(
            viewport={'width': 1280, 'height': 800},
            locale='en-US',
            extra_http_headers={'Accept-Language': 'en-US,en;q=0.9'}
        )
        page = context.new_page()
        
        def click_with_tolerance(locator, timeout=4000, retries=2):
            for _ in range(retries):
                try:
                    locator.wait_for(state="visible", timeout=timeout)
                except Exception: pass
                try:
                    locator.click()
                    return True
                except Exception:
                    try:
                        locator.click(force=True)
                        return True
                    except Exception:
                        page.wait_for_timeout(500)
            return False
        
        result = {'variants': [], 'current_details': {}}
        
        def handle_response(response):
            if 'twisterDimensionSlotsDefault' in response.url:
                try:
                    text = response.text()
                    blocks = text.split('&&&')
                    for block in blocks:
                        if 'ASIN' in block:
                            asin_match = re.search(r'"ASIN"\s*:\s*"([^"]+)"', block)
                            if not asin_match: continue
                            asin = asin_match.group(1)
                            
                            price_text_match = re.search(
                                r'"twisterSlotDiv"[^>]*>\s*<span[^>]*>\s*([^<]+)\s*<', 
                                block, re.DOTALL
                            )
                            if price_text_match:
                                price_text = price_text_match.group(1).strip()
                            else:
                                price_text_match = re.search(r'([0-9]+个选项，起始价：CNY [0-9]+\.[0-9]+)', block)
                                price_text = price_text_match.group(1) if price_text_match else None
                            
                            if not any(v['asin'] == asin for v in result['variants']):
                                result['variants'].append({'asin': asin, 'price_text': price_text})
                except Exception: pass
        
        page.on('response', handle_response)
        print(f"正在访问 (Playwright): {product_url}")
        page.goto(product_url)
        
        try:
            cont = page.get_by_role("button", name=re.compile(r"(Continue shopping|继续购物)", re.I))
            if cont.count() > 0 and click_with_tolerance(cont.first):
                page.wait_for_load_state("networkidle")
            else:
                alt = page.locator("text=Continue shopping")
                if alt.count() > 0 and click_with_tolerance(alt.first):
                    page.wait_for_load_state("networkidle")
        except Exception: pass
        
        page.wait_for_timeout(5000)
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(2000)
        
        # 同时提取当前页面的详情，减少一次 requests
        result['current_details'] = extract_details_from_html(page.content())
        browser.close()
        return result

def get_details_with_requests(asin):
    """
    使用 requests 获取商品详情。
    参考 ProductData.py 的简单请求方式，不设置复杂的 User-Agent。
    """
    url = f'https://www.amazon.com/dp/{asin}?th=1&language=en_US'
    print(f"  正在获取详情 (Requests): {asin}")
    try:
        response = requests.get(
            url,
            timeout=10,
            headers={'Accept-Language': 'en-US,en;q=0.9'}
        )
        if response.status_code == 200:
            html_content = response.text
            # 检查是否被拦截
            if "To discuss automated access to Amazon data" in html_content or "api-services-support@amazon.com" in html_content:
                print(f"  ⚠️ {asin} 被 Requests 拦截（机器人验证）。")
                return None
            return extract_details_from_html(html_content)
    except Exception as e:
        print(f"  ❌ Requests 错误 ({asin}): {e}")
    return None

def run_batch(input_path, output_excel_path):
    input_asins = read_asins_from_file(input_path)
    processed_asins = set()
    
    for base_asin in input_asins:
        if base_asin in processed_asins: continue
        
        print(f"\n--- 处理 ASIN 组: {base_asin} ---")
        url = f'https://www.amazon.com/dp/{base_asin}?th=1&language=en_US'
        # 1. 用 Playwright 拿变体列表（这个必须用 Playwright，因为涉及 AJAX 监听）
        res = capture_variants_only(url)
        if not res: continue
        
        variants = res.get('variants', [])
        folder_path = os.path.join(base_dir, base_asin)
        try:
            os.makedirs(folder_path, exist_ok=True)
        except Exception:
            pass
        group_rows = []
        
        # 2. 遍历变体获取详情
        if not variants:
            # 没变体，存基础 ASIN
            row = {'asin': base_asin, '价格文本': '未抓取到'}
            row.update(res['current_details'])
            group_rows.append(row)
            processed_asins.add(base_asin)
        else:
            for v in variants:
                v_asin = v['asin']
                if v_asin in processed_asins: continue
                
                # 如果是当前 Playwright 打开的页面，直接用已有的详情
                if base_asin == v_asin or v_asin in url:
                    details = res['current_details']
                else:
                    # 否则优先尝试 requests (速度快)
                    details = get_details_with_requests(v_asin)
                    if not details:
                        # ⚠️ 容错：如果 requests 被封，切换到 Playwright 兜底获取详情
                        print(f"  🔄 正在为 {v_asin} 启动 Playwright 兜底...")
                        v_url = f'https://www.amazon.com/dp/{v_asin}?th=1'
                        v_res = capture_variants_only(v_url)
                        if v_res:
                            details = v_res['current_details']
                        else:
                            print(f"  ⏭️ 跳过变体 {v_asin} (所有方式均获取失败)")
                            continue
                
                row = {'asin': v_asin, '价格文本': v['price_text']}
                row.update(details)
                group_rows.append(row)
                processed_asins.add(v_asin)
    
        if group_rows:
            df = pd.DataFrame(group_rows)
            def _ensure_list(x):
                if isinstance(x, list):
                    return x
                return _split_image_urls(x)
            if '商品图片' in df.columns:
                max_img = max((len(_ensure_list(v)) for v in df['商品图片']), default=0)
                for i in range(1, max_img + 1):
                    df[f'商品图片{i}'] = [(_ensure_list(v)[i-1] if len(_ensure_list(v)) >= i else '') for v in df['商品图片']]
            if '详情图片' in df.columns:
                max_xq = max((len(_ensure_list(v)) for v in df['详情图片']), default=0)
                for i in range(1, max_xq + 1):
                    df[f'详情图片{i}'] = [(_ensure_list(v)[i-1] if len(_ensure_list(v)) >= i else '') for v in df['详情图片']]
            if '商品图片' in df.columns:
                df.drop(columns=['商品图片'], inplace=True)
            if '详情图片' in df.columns:
                df.drop(columns=['详情图片'], inplace=True)
            group_excel = os.path.join(folder_path, 'asin_results.xlsx')
            try:
                writer = pd.ExcelWriter(group_excel, engine='xlsxwriter')
                display_df = df.copy()
                for c in list(display_df.columns):
                    if str(c).startswith('商品图片') or str(c).startswith('详情图片'):
                        display_df[c] = ''
                display_df.to_excel(writer, index=False, sheet_name='Sheet1')
                worksheet = writer.sheets['Sheet1']
                columns = list(display_df.columns)
                base_row_height = 120
                base_col_width = 30
                image_cols = [i for i, name in enumerate(columns) if str(name).startswith('商品图片') or str(name).startswith('详情图片')]
                for r in range(len(df)):
                    row_has_img = False
                    for col_idx in image_cols:
                        val = df.iloc[r, col_idx]
                        url_i = str(val).strip()
                        if not url_i:
                            continue
                        try:
                            resp = requests.get(url_i, timeout=10, headers={
                                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36',
                                'Accept-Language': 'en-US,en;q=0.9',
                                'Referer': 'https://www.amazon.com/'
                            })
                            if resp.status_code == 200 and resp.content:
                                img_data = BytesIO(resp.content)
                                worksheet.set_column(col_idx, col_idx, base_col_width)
                                worksheet.insert_image(r + 1, col_idx, url_i, {
                                    'image_data': img_data,
                                    'x_scale': 0.6,
                                    'y_scale': 0.6
                                })
                                row_has_img = True
                        except Exception:
                            pass
                    if row_has_img:
                        worksheet.set_row(r + 1, base_row_height)
                writer.close()
                print(f'已写入Excel: {group_excel}')
            except Exception as e:
                try:
                    writer = pd.ExcelWriter(group_excel, engine='openpyxl')
                    display_df = df.copy()
                    for c in list(display_df.columns):
                        if str(c).startswith('商品图片') or str(c).startswith('详情图片'):
                            display_df[c] = ''
                    display_df.to_excel(writer, index=False, sheet_name='Sheet1')
                    writer.close()
                    if load_workbook and XLImage and get_column_letter and PILImage:
                        wb = load_workbook(group_excel)
                        ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
                        columns = list(display_df.columns)
                        base_row_height = 120
                        base_col_width = 30
                        image_cols = [i for i, name in enumerate(columns) if str(name).startswith('商品图片') or str(name).startswith('详情图片')]
                        for r in range(len(df)):
                            row_has_img = False
                            for col_idx in image_cols:
                                val = df.iloc[r, col_idx]
                                url_i = str(val).strip()
                                if not url_i:
                                    continue
                                try:
                                    resp = requests.get(url_i, timeout=10)
                                    if resp.status_code == 200 and resp.content:
                                        pil_img = PILImage.open(BytesIO(resp.content))
                                        img = XLImage(pil_img)
                                        row = r + 2
                                        col_letter = get_column_letter(col_idx + 1)
                                        ws.column_dimensions[col_letter].width = base_col_width
                                        ws.add_image(img, f'{col_letter}{row}')
                                        row_has_img = True
                                except Exception:
                                    pass
                            if row_has_img:
                                ws.row_dimensions[r + 2].height = base_row_height
                        wb.save(group_excel)
                    print(f'已写入Excel: {group_excel}')
                except Exception as e2:
                    try:
                        df.to_excel(group_excel, index=False)
                        print(f'已写入Excel: {group_excel}')
                    except Exception as e3:
                        print(f'写入Excel失败: {e3}')

def read_asins_from_file(path):
    with open(path, 'r', encoding='utf-8') as f:
        text = f.read()
    tokens = re.split(r'\s+', text.strip())
    return [t for t in tokens if t]

import os
import sys

# 获取程序运行的根目录
if getattr(sys, 'frozen', False):
    # 打包后的 exe 运行目录
    base_dir = os.path.dirname(sys.executable)
else:
    # 源代码运行目录
    base_dir = os.path.dirname(os.path.abspath(__file__))

if __name__ == "__main__":
    try:
        os.environ["PLAYWRIGHT_BROWSERS_PATH"] = "0"
        input_path = os.path.join(base_dir, '亚马逊Asin.txt')
        output_excel = os.path.join(base_dir, 'asin_results.xlsx')
        
        if not os.path.exists(input_path):
            print(f"❌ 错误: 找不到输入文件 '{input_path}'")
            print("请确保 '亚马逊Asin.txt' 与程序在同一目录下。")
        else:
            run_batch(input_path, output_excel)
            
    except Exception as e:
        print(f"\n❌ 程序运行发生严重错误: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print("\n" + "="*30)
        input("程序运行结束，按回车键退出...")
